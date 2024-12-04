import openpyxl
from openpyxl.styles import Font, Border, Side

def read_excel_file(file_name):
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    data = []
    for row in ws.iter_rows(min_row=1, max_col=3, max_row=3, values_only=True):
        data.append(list(row))
    return data

data1 = read_excel_file('1111.xlsx')
data2 = read_excel_file('2222.xlsx')
data3 = read_excel_file('3333.xlsx')

combined_data = data1 + data2 + data3

sorted_data = sorted(combined_data, key=lambda x: (x[0], x[1], x[2]), reverse=True)

def save_sorted_data_to_excel(data, file_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SortedData'

    font = Font(bold=True, color="0000FF")
    border = Border(left=Side(style="thin", color="000000"),
                    right=Side(style="thin", color="000000"),
                    top=Side(style="thin", color="000000"),
                    bottom=Side(style="thin", color="000000"))

    for row_index, row in enumerate(data, 1):
        for col_index, value in enumerate(row, 1):
            cell = ws.cell(row=row_index, column=col_index, value=value)
            cell.font = font
            cell.border = border

    wb.save(file_name)

save_sorted_data_to_excel(sorted_data, 'sorted_output.xlsx')
